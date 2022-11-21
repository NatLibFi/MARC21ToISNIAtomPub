import io
import re
import os.path                                             
import openpyxl
from datetime import date, datetime

class RaportWriter():
    def __init__(self, file_name):
        self.output_path = file_name + date.today().isoformat() + ".xlsx"
        if os.path.exists(self.output_path):
            self.output_path = file_name + datetime.today().replace(microsecond=0).isoformat() + ".xlsx"
            self.output_path = self.output_path.replace(":", "")
        headers = ['Asteri-id', 'nimi', '024', 'ISNI reason', 'oma merkintä', 'ISNI-tunnukset']
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'epäonnistuneet'
        for idx, header in enumerate(headers):
            ws.cell(row=1, column=idx+1, value=header)
            
        headers = ['Asteri-id', 'uusi ISNI']    
        wb.create_sheet("ISNIt") 
        ws = wb['ISNIt']
        for idx, header in enumerate(headers):
            ws.cell(row=1, column=idx+1, value=header)
            
        wb.save(filename = self.output_path)
          
    def handle_response(self, response, id, record_data):
        cells = []
        cells.append(id)
        name = ""
        name_types = ['personalName', 'organisationName']
        for nt in name_types:
            if nt in record_data:
                for key in record_data[nt]:
                    if key != 'nameUse':
                        values = record_data[nt][key]
                        if values:
                            if type(values) == str:
                                name += values + " "
                            else:
                                for value in values:
                                    name += value + " "
        cells.append(name.strip())
        if 'ISNI' in record_data:
            cells.append(record_data['ISNI'])
        else:
            cells.append('')
        if response.get('isni'):
            if record_data.get('ISNI'):
                if response['isni'].replace(' ', '') != record_data['ISNI']:
                    cells.append('eri ISNI')
                    cells.append('')
                    cells.append(response['isni'])
                    self.write_response('epäonnistuneet', cells)
            else:
                del cells[-1]
                cells.append(response['isni'])
                self.write_response('ISNIt', cells)
        elif response.get('possible matches'):
            if response.get('reason'):
                cells.append(response['reason'])
            else:
                cells.append('')
            cells.append('')
            for pm in response['possible matches']:
                cells.append(pm['id'])
                if 'evaluation score' in pm:
                    cells.append(pm['evaluation score'])
            self.write_response('epäonnistuneet', cells)
        elif response.get('error'):
            cells.extend(response['error'])
            self.write_response('epäonnistuneet', cells)
        elif response.get('reason'):
            cells.append(response['reason'])
            self.write_response('epäonnistuneet', cells)
            
    def write_response(self, work_sheet_name, cells):
        wb = openpyxl.load_workbook(self.output_path)
        ws = wb[work_sheet_name]
        row_number = ws.max_row + 1
        for idx, value in enumerate(cells):
            ws.cell(row=row_number, column=idx+1, value=value)
        wb.save(filename = self.output_path)
        