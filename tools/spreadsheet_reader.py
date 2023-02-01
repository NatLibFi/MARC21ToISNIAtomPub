import csv
import logging
import openpyxl
import os.path
import sys

class SpreadsheetReader:

    def __init__(self, file_path):
        file_extension = os.path.splitext(file_path)
        file_format = None
        if len(file_extension) > 1:
            file_format = file_extension[1][1:]
        if not file_format or file_format not in ['csv', 'xlsx']:
            logging.error("File extension should be csv, xlsx or xls")
            sys.exit(2)
        try:
            self.reader = CSVReader(file_path)
        except UnicodeDecodeError:
            self.reader = ExcelReader(file_path)

        self.header_indices = self.reader.header_indices

    def close(self) -> None:
        self.reader.close()

    def __next__(self):
        row = next(self.reader)
        if row == None:
            raise StopIteration
        return row

    def __iter__(self):
        return self

class CSVReader(SpreadsheetReader):
    def __init__(self, file_path):
        sniffer = csv.Sniffer()
        with open(file_path) as fh:
            delimiter = sniffer.sniff(fh.read(5000)).delimiter
        encodings = ['utf-8', 'utf-8-sig']
        for encoding in encodings:
            self.fh = open(file_path, mode='r', encoding=encoding, newline='')
            self.reader = csv.reader(self.fh, delimiter=delimiter)
            try:
                header = next(self.reader)
            except UnicodeDecodeError:
                logging.error("Encoding of CSV file %s not supported"%file_path)
                sys.exit(2)
            if header[0].startswith('\ufeff'):
                continue
            else:
                break
        self.header_indices = {value:n for n, value in enumerate(header)}
    
    def close(self) -> None:
        self.fh.close()

    def __next__(self):
        return next(self.reader)

class ExcelReader(SpreadsheetReader):
    def __init__(self, file_path):
        self.fh = openpyxl.load_workbook(file_path)
        self.reader = self.fh.active
        self.max_row = self.reader.max_row
        self.row_number = 2 
        for row in self.reader.iter_rows(min_row=1, max_row=1):
            self.header_indices = {cell.value:n for n, cell in enumerate(row)} 

    def close(self) -> None:
        self.fh.close()

    def __next__(self):
        for row in self.reader.iter_rows(min_row=self.row_number, max_row=self.row_number):
            row = [cell.value for cell in row]
        if self.row_number == self.max_row:
            row = None
        self.row_number += 1
        return row
