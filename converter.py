import logging
from lxml import etree
import os
import io
import re
import sys
import openpyxl
import argparse
import configparser
import requests
from datetime import datetime
from isni_request import create_xml
from marc21_converter import MARC21Converter
from gramex_converter import GramexConverter
from tools import parse_isni_response
from tools import xlsx_raport_writer
from tools import api_query

class Converter():
    """
    A class that transform files to ISNI Atom Pub XML format.
    """
    
    def __init__(self):
        parser = argparse.ArgumentParser(description="ISNI AtomPub Conversion program")
        parser.add_argument("-f", "--format",
            help="File format", choices=['marc21', 'alephseq', 'gramex'], required=True)
        parser.add_argument("-a", "--authority_files",
            help="File containing identity data")
        parser.add_argument("-r", "--resource_files", 
            help="File containing titles of works of authors")
        parser.add_argument("-d", "--output_directory",
            help="Output directory for ISNI AtomPub XML files")
        parser.add_argument("-v", "--validation_file",
            help="Enter file path of ISNI Atom Pub Request XSD file to validate XML requests")
        parser.add_argument("-i", "--identifier",
            help="Identifier of the database of requestor, e. g. FI-ASTERI-N")
        parser.add_argument("-o", "--origin",
            help="Source code of origin, e. g. NLFIN")
        parser.add_argument("-c", "--concat", action='store_true',
            help="Concatenate XML request into one file")
        parser.add_argument("-D", "--dirmax", type=int,
            help="Number of output XML files in one directory, default 100")
        parser.add_argument("-t", "--identity_types",
            help="Restrict requested records either to persons or organisations", choices=['persons', 'organisations'])
        parser.add_argument("-u", "--until",
            help="Request records created or modified before the set date formatted YYYY-MM-DD")
        input_group = parser.add_mutually_exclusive_group()
        input_group.add_argument("-M", "--modified_after",
            help="Request records modified or created on or after the set date formatted YYYY-MM-DD")
        input_group.add_argument("-C", "--created_after",
            help="Request records created on or after the set date formatted YYYY-MM-DD")
        input_group.add_argument("-l", "--id_list",
            help="Path of text file containing local identifiers, one in every row, of records to be requested to ISNI requestor")
        input_group.add_argument("-I", "--input_raport_list",
            help="Path of CSV file containing merge instructions for ISNI requests")
        parser.add_argument("-R", "--output_raport_list",
            help="File path of CSV file raport for unsuccesful ISNI requests")
        parser.add_argument("-O", "--output_marc_fields",
            help="File path for Aleph sequential MARC21 fields code 024 where received ISNI identifiers are written along recent identifiers")
        parser.add_argument("--output_to_api",
            help="ISNI identifiers from ISNI responses are output to API, address is defined in configuration file under section AUT NAMES API")
        parser.add_argument("-m", "--mode",
            help="Mode of program: Write requests into a directory or send them to ISNI or test sending to test database", choices=['write', 'prod', 'test'], required=True)
        parser.add_argument("-F", "--config_file_path",
            help="File path for configuration file structured for Python ConfigParser")
        args = parser.parse_args()
        self.converter = None
        self.modified_after = None
        self.created_after = None
        self.config = configparser.ConfigParser()
        self.config.read(args.config_file_path)
        self.convert_to_atompub(args)

    def convert_to_atompub(self, args):
        """
        convert input data to ISNI AtomPub XML format
        :param args: command line arguments parsed by ConfigParser
        """
        logging.getLogger().setLevel(logging.INFO)
        if args.mode in ['prod', 'test']:
            if args.mode == 'prod':
                section = self.config['ISNI SRU API']
            if args.mode == 'test':
                section = self.config['ISNI SRU TEST API']
            username = None
            password = None
            if 'ISNI_USER' in os.environ and 'ISNI_PASSWORD' in os.environ:
                username = os.environ['ISNI_USER']
                password = os.environ['ISNI_PASSWORD']
            self.sru_api_query = api_query.APIQuery(config_section=section,
                                        username=username,
                                        password=password)
        if args.modified_after:
            self.modified_after = datetime.date(datetime.strptime(args.modified_after, "%Y-%m-%d"))
        if args.created_after:
            self.created_after = datetime.date(datetime.strptime(args.created_after, "%Y-%m-%d"))
        
        if args.mode == 'write':
            if not args.output_directory:
                logging.error("Mode is set to write. Set output directory parameter for output files")
                sys.exit(2)
            
        if (args.authority_files or args.resource_files) and not args.format:
            logging.error("Using authority_files or resource_files command line arguments but argument format is missing.")
            sys.exit(2)

        logging.info("Conversion started: %s"%datetime.now().replace(microsecond=0).isoformat())
        requested_ids = set()
        if args.id_list:
            with open(args.id_list, 'r', encoding='utf-8') as fh:
                for row in fh:
                    requested_ids.add(row.rstrip())
        merge_instructions = {}
        if args.input_raport_list:
            wb = openpyxl.load_workbook(args.input_raport_list)
            ws = wb.active
            
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
                local_id = None
                if row[0].value:
                    local_id = str(row[0].value).strip()
                else:
                    logging.error("Local id missing in row number %s"%idx)
                if local_id:
                    if row[4].value:
                        if row[4].value in ['merge', 'isNot', 'resend']:
                            merge_instructions[local_id] = {'instruction': row[4].value, 'identifiers': []}
                        else:
                            logging.error("Incorrect instruction %s in row number %s"%(row[4].value, row[0].row))
                    if local_id in merge_instructions:
                        for col in ws.iter_cols(min_row=row[0].row, max_row=row[0].row, min_col=6, max_col=ws.max_column):
                            if col[0].value:
                                isni_identifier = self.validate_isni_id(str(col[0].value).strip())
                                merge_instructions[local_id]['identifiers'].append(isni_identifier)
                        requested_ids.add(local_id)

        if args.format in ['marc21', 'alephseq']:
            self.converter = MARC21Converter(self.config)
        elif args.format == 'gramex':
            self.converter = GramexConverter(self.config)
        records = self.converter.get_authority_data(args, requested_ids)
        concat = False
        xmlschema = None
        dirmax = 100
        if args.dirmax:
            dirmax = args.dirmax
        if args.validation_file:
            with open(args.validation_file, 'rb') as fh:
                xmlschema_doc = etree.parse(fh)
                xmlschema = etree.XMLSchema(xmlschema_doc)   
        if args.output_directory and not os.path.exists(args.output_directory):
            os.mkdir(args.output_directory)
        dirindex = 0
        xml = ""
        if args.concat:
            concat_path = os.path.join(args.output_directory, "concat.xml")
            with open(concat_path, 'ab+') as concat_file:
                concat_file.write(bytes("<?xml version=\"1.0\" ?>\n<root>\n", "UTF-8"))
        idx = 0
        logging.info("Starting to convert records...")

        isnis = {}
        if args.output_raport_list:
            raport_writer = xlsx_raport_writer.RaportWriter(args.output_raport_list)
        for record_id in records:
            merge_instruction = None
            merge_identifiers = []
            if merge_instructions:
                if record_id not in merge_instructions:
                    continue
                else:
                    merge_instruction = merge_instructions[record_id]['instruction']
                    merge_identifiers = merge_instructions[record_id]['identifiers']
            xml = None
            if not records[record_id]['errors']:
                xml = create_xml(records[record_id], merge_instruction, merge_identifiers)
            if idx % dirmax == 0:
                dirindex += 1
            if args.output_directory:
                subdir = os.path.join(args.output_directory, str(dirindex).zfill(3))
                if not (os.path.exists(subdir)) and not concat:
                    os.mkdir(subdir)
            if xml and xmlschema:
                if not self.valid_xml(record_id, xml, xmlschema):
                    continue

            if args.mode in ["prod", "test"]:
                isni_data = {'errors': []}
                if xml:
                    logging.info("Sending record %s"%record_id)
                    response = self.send_xml(xml, args.mode, args.origin)
                    if response.status_code != 200:
                        isni_data['errors'].extend([line for line in response.text.splitlines() if line])
                    else:
                        isni_data.update(parse_isni_response.dictify_xml(response.text)[0])
                    # if record is not entered in ISNI database, but has possible matches:
                    if 'possible matches' in isni_data and not 'ppn' in isni_data and not 'isni' in isni_data:
                        for pm in isni_data['possible matches']:
                            if 'ppn' in pm:
                                ppn = pm['ppn']
                                result = self.sru_api_query.get_isni_query_data('ppn='+ppn)
                                if result:
                                    result = result[0]
                                    if 'isni' in result:
                                        pm['isni'] = result['isni']
                                    source_ids = result['sources']
                                    pm['sources'] = {code: re.sub("[\(].*?[\)]", "", source_ids[code]) for code in source_ids}
                            else:
                                isni_data['errors'].append('Record has possible match in ISNI without id')
                isni_data['errors'].extend(records[record_id]['errors'])
                isnis[record_id] = isni_data
                if args.output_raport_list:
                    raport_writer.handle_response(isni_data, record_id, records[record_id])
            elif args.mode == "write":
                if xml:
                    if args.concat:
                        xml_path = concat_path
                    else:
                        xml_path = os.path.join(subdir, record_id + ".xml")
                    self.write_xml(xml, xml_path, args.concat)
            idx += 1
        logging.info("Conversion done for %s items"%idx)
        if args.concat:
            with open(args.output_directory+"/concat.xml", 'ab+') as concat_file:
                concat_file.write(bytes("</root>", "UTF-8"))
        if isnis:
            isni_records = self.converter.create_isni_fields(isnis, args.identifier)
            if args.output_marc_fields:
                self.converter.write_isni_fields(args.output_marc_fields, isni_records)

    def valid_xml(self, record_id, xml, xmlschema):
        try:
            string_xml = io.StringIO(xml)
            doc = etree.parse(string_xml)
            xmlschema.assert_(doc)
            return True
        except AssertionError as e:
            logging.error("Record %s XML AssertionError : %s"%(record_id, e))

    def validate_isni_id(self, isni_id):
        """Validate ISNI identifier in case of typos"""
        isni_id = isni_id.replace(' ', '')
        if len(isni_id) == 16:
            return {'identifier': isni_id, 'type': 'ISNI'}
        elif len(isni_id) == 9:
            return {'identifier': isni_id, 'type': 'PPN'}
        else:
            logging.error('The length of ISNI identifier %s is not 9 or 16 characters'%isni_id)

    def write_xml(self, xml, file_path, concat):
        """
        Write XML records to file. Creates directory dirname, and creates XML Request files into the folder.
        :param dirname: path of the directory, where XML records are written.
        :param dirmax: dirmax default is 100, so it makes 100 request xml files per folder before creating a new one.
        :param dirindex: index number of directory
        """ 
        xml = xml.replace("<?xml version=\"1.0\" ?>", "")
        if concat:
            xmlfile = open(file_path, 'ab+')
            xmlfile.write(bytes(xml, 'UTF-8'))
            xmlfile.flush()
        else:
            xmlfile = open(file_path, 'wb+')
            xmlfile.write(bytes(xml, 'UTF-8'))
            xmlfile.close()

    def send_xml(self, xml, mode, origin=""):
        """
        :param xml: string converted XML elementtree in ISNI AtomPub format
        :param section: section of config file, section should contain ISNI production and accept database baseurl
        :param mode: 'prod' or 'test' to choose between ISNI production and accept database
        :param origin: source code for AtomPub records
        """
        headers = {'Content-Type': 'application/atom+xml; charset=utf-8'}
        if mode == 'prod':
            section = self.config['ISNI ATOMPUB API']
        elif mode == 'test':
            section = self.config['ISNI ATOMPUB TEST API']
        url = section.get('baseurl')
        if origin:
            url += 'ORIGIN=' + origin
        response = requests.post(url, data=xml.encode('utf-8'), headers=headers)

        return response

if __name__ == '__main__':
    Converter()